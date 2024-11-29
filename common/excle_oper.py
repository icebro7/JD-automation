import os
from openpyxl.reader.excel import load_workbook


def get_excel_path():
    print(os.path.abspath(os.path.dirname(__file__)).split("common")[0])
    return os.path.abspath(os.path.dirname(__file__)).split("common")[0]
    # 获得文件路径前半段


class ExcelOper:

    def read_excel(self):
        excel_path = os.path.join(get_excel_path(), "data", "login.xlsx")
        # 打开 Excel 文件
        workbook = load_workbook(excel_path)
        sheet = workbook['Sheet1']
        # 选择要读取的工作表，这里假设工作表名为 'Sheet1'

        all_list = []
        for rows in range(2, sheet.max_row + 1):
            # 取出行数据，过滤表头
            temp_list = []
            for cols in range(1, sheet.max_column):
                temp_list.append(sheet.cell(rows, cols).value)
            all_list.append(temp_list)
        return all_list


if __name__ == '__main__':
    get_excel_path()
